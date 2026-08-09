"""AI model client for tender generation.

Given the compiled project context, the workflow-step prompts and the project
documents, this produces two artefacts: a **pricing sheet** (.xlsx) and a
**tender document**.

Connection details (provider, API key, endpoint / model id) are supplied via
environment variables — see app/settings.py. Until they are set,
``generate_tender`` returns clearly-labelled **placeholder** outputs so the
end-to-end flow (and the download UI) works today; swap in the real model by
filling the env vars and completing the provider branch in ``_invoke_remote``.

Return shape:
    {
      "ai_used": bool,
      "notes": str,
      "pricing": {"filename": str, "bytes": bytes, "media_type": str},
      "tender":  {"filename": str, "bytes": bytes, "media_type": str},
    }
"""

from __future__ import annotations

import datetime as _dt
import io
import logging

from app import settings

LOGGER = logging.getLogger(__name__)

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MD_MEDIA = "text/markdown"


class AIClientError(RuntimeError):
    """Raised when a configured model call cannot be completed."""


def _safe(name: str) -> str:
    keep = "".join(c if (c.isalnum() or c in " -_()") else "_" for c in name)
    return keep.strip() or "project"


# --------------------------------------------------------------------------- #
# Draft pricing items (step 1 of the two-pass flow)                            #
# --------------------------------------------------------------------------- #

def generate_draft_items(export: dict, context_markdown: str,
                         prompts: list[dict], documents: list[dict]) -> dict:
    """Produce the DRAFT pricing line items the estimator reviews one by one.

    With a configured model this should ask the model to derive line items
    from the documents + context; until then we derive placeholder items
    deterministically from the answered context so the review loop works.
    Returns {"ai_used": bool, "items": [{group,item,detail,qty,unit}...]}.
    """
    if settings.ai_configured():
        # TODO: call the model (same providers as _invoke_remote) with a
        # "derive line items" instruction and parse its JSON response.
        # Fall through to the deterministic derivation until implemented.
        pass

    items: list[dict] = []
    for c in export.get("context", []):
        ans = c.get("answer")
        if isinstance(ans, (list, tuple)):
            ans = ", ".join(str(x) for x in ans)
        items.append({
            "group": c.get("group", ""),
            "item": c.get("subelement", "") or c.get("question", ""),
            "detail": c.get("question", ""),
            "qty": str(ans),
            "unit": c.get("unit", ""),
        })
    if not items:
        # No context answered — seed a minimal generic take-off so the
        # review loop still functions.
        items = [
            {"group": "General", "item": "Main roof waterproofing",
             "detail": "Field area waterproofing system", "qty": "", "unit": "m²"},
            {"group": "General", "item": "Perimeter details",
             "detail": "Upstands, edge trims and terminations", "qty": "", "unit": "lm"},
            {"group": "General", "item": "Preliminaries & access",
             "detail": "Scaffold/access, welfare, supervision", "qty": "", "unit": "item"},
        ]
    notes = ""
    cp = export.get("client_pricing_sheet") or {}
    if cp.get("enabled"):
        notes = (f"Client pricing sheet flagged ('{cp.get('filename') or 'not selected'}') — "
                 "final output will populate that sheet.")
    return {"ai_used": False, "items": items, "notes": notes}


# --------------------------------------------------------------------------- #
# Placeholder artefacts (used until a real model is wired up)                  #
# --------------------------------------------------------------------------- #

def _placeholder_pricing(export: dict) -> bytes:
    """A real .xlsx pricing sheet skeleton seeded from the project context.
    Prices are left blank for the (future) AI model / estimator to populate;
    markup and waste % are surfaced so the structure is ready to compute."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    proj = export.get("project", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Sheet"

    head = Font(bold=True, color="FFFFFF")
    headfill = PatternFill("solid", fgColor="1F3864")

    ws["A1"] = f"Pricing Sheet — {proj.get('name', '')}"
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A2"] = ("PLACEHOLDER — AI model not configured. Generated from project "
                "context; rates to be populated.")
    ws["A2"].font = Font(italic=True, color="C00000")
    cp = export.get("client_pricing_sheet") or {}
    if cp.get("enabled"):
        ws["A3"] = (f"CLIENT PRICING SHEET FLAGGED: '{cp.get('filename') or 'not selected'}' — "
                    "the final estimate must populate that uploaded sheet "
                    "(add columns/values to it), not this workbook.")
        ws["A3"].font = Font(italic=True, bold=True, color="854F0B")

    ws["A4"] = "Client"; ws["B4"] = proj.get("client", "")
    ws["A5"] = "Reference"; ws["B5"] = proj.get("reference", "")
    ws["A6"] = "Profit markup (%)"; ws["B6"] = proj.get("markup_pct")
    ws["A7"] = "Waste factor (%)"; ws["B7"] = proj.get("waste_pct")
    for r in range(4, 8):
        ws[f"A{r}"].font = Font(bold=True)

    headers = ["Element Group", "Item", "Detail", "Qty", "Unit",
               "Unit rate (£)", "Material (£)", "Labour (£)", "Line total (£)",
               "Qualifications"]
    hr = 9
    for i, h in enumerate(headers, start=1):
        c = ws.cell(hr, i, h)
        c.font = head
        c.fill = headfill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    draft_items = export.get("draft_items") or []
    r = hr + 1
    if draft_items:
        for it in draft_items:
            ws.cell(r, 1, it.get("group", ""))
            ws.cell(r, 2, it.get("item", ""))
            ws.cell(r, 3, it.get("detail", ""))
            ws.cell(r, 4, it.get("qty", ""))
            ws.cell(r, 5, it.get("unit", ""))
            q = it.get("qualification", "")
            ws.cell(r, 10, q if q else ("(skipped)" if it.get("skipped") else ""))
            r += 1
    else:
        for item in export.get("context", []):
            ans = item.get("answer")
            if isinstance(ans, (list, tuple)):
                ans = ", ".join(str(x) for x in ans)
            ws.cell(r, 1, item.get("group", ""))
            ws.cell(r, 2, item.get("subelement", ""))
            ws.cell(r, 3, item.get("question", ""))
            ws.cell(r, 4, str(ans))
            ws.cell(r, 5, item.get("unit", ""))
            r += 1
        if r == hr + 1:
            ws.cell(r, 1, "(No context answers captured yet.)")

    widths = [22, 26, 40, 12, 8, 12, 12, 12, 13, 40]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _placeholder_tender(export: dict, context_markdown: str,
                        prompts: list[dict]) -> bytes:
    proj = export.get("project", {})
    steps = "\n".join(f"  {i+1}. {p['name']}" for i, p in enumerate(prompts)) or "  (none found)"
    body = f"""# Tender — {proj.get('name', '')}

*Placeholder tender. The AI model is not yet configured; once connection
details are supplied this document will be produced by the model from the
project context and the workflow steps below.*

**Client:** {proj.get('client', '')}
**Reference:** {proj.get('reference', '')}
**Client pricing sheet:** {(export.get('client_pricing_sheet') or {}).get('filename') or 'none flagged'}
**Labour rates source:** {_labour_line(export)}
**Profit markup:** {proj.get('markup_pct')}%
**Waste factor:** {proj.get('waste_pct')}%
**Generated:** {_dt.datetime.now().strftime('%d %b %Y, %H:%M')}

## Workflow steps that will be executed
{steps}

## Item qualifications
{_quals_section(export)}

## Compiled project context
{context_markdown}
"""
    return body.encode("utf-8")


def _labour_line(export: dict) -> str:
    lab = export.get("labour_rates") or {}
    src, fn = lab.get("source"), lab.get("filename", "")
    if src == "project":
        return f"project upload ({fn})"
    if src == "library":
        return f"shared library — most recent ({fn})"
    return "none available (internal assumptions)"


def _quals_section(export: dict) -> str:
    items = export.get("draft_items") or []
    quals = [it for it in items if it.get("qualification")]
    if not items:
        return "_Draft pricing has not been reviewed item by item._"
    if not quals:
        return "_No qualifications were added during the item review._"
    lines = []
    for it in quals:
        lines.append(f"- **{it.get('item','')}** ({it.get('group','')}): "
                     f"{it['qualification']}")
    return "\n".join(lines)


def _placeholder_result(export: dict, context_markdown: str,
                        prompts: list[dict], reason: str) -> dict:
    name = _safe(export.get("project", {}).get("name", "project"))
    return {
        "ai_used": False,
        "notes": reason,
        "pricing": {
            "filename": f"Pricing Sheet {name}.xlsx",
            "bytes": _placeholder_pricing(export),
            "media_type": XLSX_MEDIA,
        },
        "tender": {
            "filename": f"Tender {name}.md",
            "bytes": _placeholder_tender(export, context_markdown, prompts),
            "media_type": MD_MEDIA,
        },
    }


# --------------------------------------------------------------------------- #
# Real model call (scaffold — complete when connection details are available) #
# --------------------------------------------------------------------------- #

def _invoke_remote(export: dict, context_markdown: str,
                   prompts: list[dict], documents: list[dict]) -> dict:
    """Call the configured model and return the result dict.

    The payload assembled here is provider-agnostic; wire the actual request
    in the branch matching AI_PROVIDER. Each branch must return a result dict
    in the shape documented at the top of this module.
    """
    provider = (settings.AI_PROVIDER or "").strip().lower()

    payload = {
        "project": export.get("project", {}),
        "context_markdown": context_markdown,
        "workflow_steps": [{"name": p["name"], "text": p["text"]} for p in prompts],
        "documents": [{"filename": d["filename"]} for d in documents],
        # When enabled, the model must populate THIS uploaded sheet (add
        # columns/values) instead of producing a fresh pricing workbook.
        "client_pricing_sheet": export.get("client_pricing_sheet")
                                 or {"enabled": False, "filename": ""},
        # Which labour-rates document applies (project upload, shared library
        # fallback, or none). The shared doc itself is in `documents`, its
        # filename prefixed "[Shared labour rates] ".
        "labour_rates": export.get("labour_rates")
                         or {"source": "none", "filename": ""},
    }

    if provider == "http":
        import json
        import requests  # lazy import

        resp = requests.post(
            settings.AI_ENDPOINT,
            headers={
                "Authorization": f"Bearer {settings.AI_API_KEY}",
                "Content-Type": "application/json",
            },
            data=json.dumps(payload),
            timeout=settings.AI_TIMEOUT_SECONDS,
        )
        resp.raise_for_status()
        # Expected contract: JSON with base64-encoded files. Adjust to match the
        # real endpoint once known.
        import base64
        data = resp.json()
        name = _safe(export.get("project", {}).get("name", "project"))
        return {
            "ai_used": True,
            "notes": "Generated by configured AI model (http).",
            "pricing": {
                "filename": data.get("pricing_filename", f"Pricing Sheet {name}.xlsx"),
                "bytes": base64.b64decode(data["pricing_b64"]),
                "media_type": data.get("pricing_media_type", XLSX_MEDIA),
            },
            "tender": {
                "filename": data.get("tender_filename", f"Tender {name}.docx"),
                "bytes": base64.b64decode(data["tender_b64"]),
                "media_type": data.get(
                    "tender_media_type",
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document"),
            },
        }

    if provider == "bedrock":
        # TODO: implement the AWS Bedrock call, e.g.:
        #   import boto3
        #   rt = boto3.client("bedrock-runtime", region_name=settings.AWS_REGION)
        #   out = rt.invoke_model(modelId=settings.AI_MODEL_ID, body=json.dumps(payload))
        # then parse `out` into the two artefacts.
        raise AIClientError(
            "Bedrock provider is selected but the bedrock call is not yet "
            "implemented in ai_client._invoke_remote (add the invoke_model call)."
        )

    raise AIClientError(f"Unknown AI_PROVIDER '{settings.AI_PROVIDER}'.")


# --------------------------------------------------------------------------- #
# Public entry point                                                          #
# --------------------------------------------------------------------------- #

def generate_tender(export: dict, context_markdown: str,
                    prompts: list[dict], documents: list[dict]) -> dict:
    """Produce the pricing sheet + tender document. Uses the configured model
    when available, otherwise returns labelled placeholder artefacts."""
    if not settings.ai_configured():
        return _placeholder_result(
            export, context_markdown, prompts,
            reason="AI model not configured — placeholder pricing sheet and "
                   "tender produced. Set AI_PROVIDER and the connection details "
                   "to generate real documents.",
        )
    return _invoke_remote(export, context_markdown, prompts, documents)
